import os
import pandas as pd
from openai import OpenAI, Completion
import logging
from openpyxl import load_workbook

logging.basicConfig(filename='metadata_generation.log', level=logging.INFO)

def generate_metadata(xlsx_file, videos_folder):
    # Validate inputs
    if not os.path.isdir(videos_folder):
        raise ValueError(f"{videos_folder} is not a valid folder")
        
    if not os.path.isfile(xlsx_file):
        raise ValueError(f"{xlsx_file} is not a valid xlsx file")
        
    openai = OpenAI(api_key="YOUR_API_KEY") 
    
    # Load existing Excel file
    wb = load_workbook(xlsx_file)
    ws = wb.active
    
    # Get list of video files    
    video_files = os.listdir(videos_folder)
    
    logging.info(f"Found {len(video_files)} video files")
    
    # Add new video filenames if not already present
    for video_filename in video_files:
        if video_filename not in [cell.value for cell in ws["A"]]:
            logging.info(f"Adding new video: {video_filename}")
            ws.append([video_filename])
            
    # Generate descriptions, keywords, and categories
    for row, cell in enumerate(ws["B"]):
        if cell.value is None:
            # Use OpenAI API to generate description based on filename
            prompt = f"Describe a video titled '{ws.cell(row=row+1, column=1).value}' in 200 characters using at least 5 words"        
            try:
                response = openai.Completion.create(engine="text-davinci-002", prompt=prompt, max_tokens=200)
            except Exception as e:
                logging.error(f"OpenAI API call failed for description: {e}")
                continue

            desc = response.choices[0].text.strip()

            if len(desc.split()) < 5 or len(desc) > 200:
                logging.warning(f"Invalid description for {ws.cell(row=row+1, column=1).value}, retrying...")
                continue

            ws.cell(row=row+1, column=2).value = desc
            logging.info(f"Added description for {ws.cell(row=row+1, column=1).value}")

            # Use OpenAI API to generate keywords based on description
            keywords_prompt = f"Based on the description: '{desc}', what could be 8-49 unique keywords for the video?"
            try:
                keywords_response = openai.Completion.create(engine="text-davinci-002", prompt=keywords_prompt, max_tokens=150)
            except Exception as e:
                logging.error(f"OpenAI API call failed for keywords: {e}")
                continue

            keywords = keywords_response.choices[0].text.strip().split(",")

            # Remove duplicate keywords for this video
            keywords = list(dict.fromkeys(keywords))

            if len(keywords) < 8 or len(keywords) > 49:
                logging.warning(f"Invalid keyword list for {ws.cell(row=row+1, column=1).value}, retrying...")
                continue

            ws.cell(row=row+1, column=3).value = ",".join(keywords)
            logging.info(f"Added {len(keywords)} keywords for {ws.cell(row=row+1, column=1).value}")

            # Use OpenAI API to generate category based on description
            category_prompt = f"Based on the description: '{desc}', what could be a suitable category for the video?"
            try:
                category_response = openai.Completion.create(engine="text-davinci-002", prompt=category_prompt, max_tokens=10)
            except Exception as e:
                logging.error(f"OpenAI API call failed for category: {e}")
                continue

            category = category_response.choices[0].text.strip()
            ws.cell(row=row+1, column=4).value = category
            logging.info(f"Added category for {ws.cell(row=row+1, column=1).value}")
    
    # Save updated Excel file
    wb.save(xlsx_file)
